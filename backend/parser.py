"""
Danish Rent Roll Parser

Parses Danish rent roll files (Excel and PDF) into a standardized format.
Includes automatic header detection, multi-sheet handling, OCR for scanned PDFs,
and Danish number format conversion.
"""

import os
import re
from typing import Optional
from pathlib import Path

import openpyxl
import xlrd
import pdfplumber
from pdf2image import convert_from_path
import pytesseract
from PIL import Image


class ParseError(Exception):
    """Custom exception for parsing errors."""

    ERROR_TYPES = {
        "file_not_found": "File not found",
        "invalid_file_type": "Invalid file type. Supported: .xlsx, .xls, .pdf",
        "password_protected": "File is password protected",
        "no_header_found": "Could not identify header row in first 15 rows",
        "no_data_found": "No data rows found after header",
        "ocr_failed": "OCR processing failed",
        "corrupted_file": "File appears to be corrupted",
    }

    def __init__(self, error_type: str, message: str = None):
        self.error_type = error_type
        self.message = message or self.ERROR_TYPES.get(error_type, "Unknown error")
        super().__init__(self.message)

    def to_dict(self) -> dict:
        return {
            "success": False,
            "error": self.error_type,
            "message": self.message,
        }


# Keywords for header detection (lowercase for matching)
HEADER_KEYWORDS = [
    # Danish
    "lejemål", "lejnr", "areal", "m2", "kvm", "leje", "årlig",
    "nr", "bel", "etage", "bem", "bemærkning", "reg", "regulering",
    "start", "indflytning", "type", "anvendelse",
    # English
    "unit_id", "size_sqm", "sqm", "area", "rent", "floor", "address",
    "zipcode", "postal", "door", "rooms", "lease", "tenant"
]

# Column mapping to standard field names (supports Danish and English)
COLUMN_MAPPING = {
    # Unit ID
    "lejnr": "unit_id",
    "lejemål": "unit_id",
    "lejenr": "unit_id",
    "lejemålsnr": "unit_id",
    "nr": "unit_id",
    "unit_id": "unit_id",
    # Square meters
    "areal": "sqm",
    "m2": "sqm",
    "kvm": "sqm",
    "m²": "sqm",
    "kvadratmeter": "sqm",
    "size_sqm": "sqm",
    "sqm": "sqm",
    "area": "sqm",
    "size": "sqm",
    # Annual rent
    "leje": "annual_rent",
    "årlig": "annual_rent",
    "årligleje": "annual_rent",
    "årlig leje": "annual_rent",
    "husleje": "annual_rent",
    "rent_current_gri": "annual_rent",
    "rent_current": "annual_rent",
    "annual_rent": "annual_rent",
    "rent": "annual_rent",
    "gri": "annual_rent",
    # Floor
    "bel": "floor",
    "etage": "floor",
    "beliggenhed": "floor",
    "unit_floor": "floor",
    "floor": "floor",
    # Unit type
    "type": "unit_type",
    "anvendelse": "unit_type",
    "lejemålstype": "unit_type",
    "unit_type": "unit_type",
    # Lease start
    "start": "lease_start",
    "indflytning": "lease_start",
    "startdato": "lease_start",
    "indflytningsdato": "lease_start",
    "lease_start": "lease_start",
    # Lease end
    "lease_end": "lease_end",
    "slutdato": "lease_end",
    "fraflytning": "lease_end",
    # Notes
    "bem": "notes",
    "bemærkning": "notes",
    "bemærkninger": "notes",
    "note": "notes",
    "noter": "notes",
    "notes": "notes",
    # Rent adjustment
    "reg": "rent_adjustment",
    "regulering": "rent_adjustment",
    "lejeregulering": "rent_adjustment",
    # Address
    "unit_address": "address",
    "address": "address",
    "adresse": "address",
    # Postal code
    "unit_zipcode": "postal_code",
    "zipcode": "postal_code",
    "postal_code": "postal_code",
    "postnr": "postal_code",
    "postnummer": "postal_code",
    # Door
    "unit_door": "door",
    "door": "door",
    "dør": "door",
    # Rooms
    "rooms_amount": "rooms",
    "rooms": "rooms",
    "værelser": "rooms",
    # Tenant
    "tenant_name1": "tenant_name",
    "tenant_name": "tenant_name",
    "lejer": "tenant_name",
    "lejernavn": "tenant_name",
    # Unit status (for vacancy detection)
    "units_status": "unit_status",
    "unit_status": "unit_status",
    "status": "unit_status",
    "lejestatus": "unit_status",
    "udlejningsstatus": "unit_status",
}

# End row indicators
END_ROW_KEYWORDS = ["total", "sum", "i alt", "ialt", "samlet", "subtotal"]

# Unit type categorization
UNIT_TYPE_BOLIG = ["apartment", "residential", "bolig", "lejlighed", "værelse", "room"]
UNIT_TYPE_ERHVERV = ["commercial", "retail", "office", "erhverv", "kontor", "butik", "lager", "warehouse"]
UNIT_TYPE_PARKERING = ["parking", "parkering", "p-plads", "garage", "carport"]

# Vacancy status keywords
VACANCY_KEYWORDS = ["vacant", "ledig", "tom", "fraflyttet", "empty", "available", "til leje"]


def categorize_unit_type(unit_type_value: str) -> str:
    """Categorize unit type into Bolig, Erhverv, Parkering, or Andet."""
    if not unit_type_value:
        return "andet"

    val_lower = unit_type_value.lower().strip()

    for keyword in UNIT_TYPE_BOLIG:
        if keyword in val_lower:
            return "bolig"

    for keyword in UNIT_TYPE_ERHVERV:
        if keyword in val_lower:
            return "erhverv"

    for keyword in UNIT_TYPE_PARKERING:
        if keyword in val_lower:
            return "parkering"

    return "andet"


def is_unit_vacant(status_value: str) -> bool:
    """Check if a unit is vacant based on status value."""
    if not status_value:
        return False

    val_lower = status_value.lower().strip()

    for keyword in VACANCY_KEYWORDS:
        if keyword in val_lower:
            return True

    return False


def detect_file_type(file_path: str) -> str:
    """Detect file type by extension."""
    ext = Path(file_path).suffix.lower()
    if ext in [".xlsx", ".xls"]:
        return "excel"
    elif ext == ".pdf":
        return "pdf"
    else:
        raise ParseError("invalid_file_type")


def convert_danish_number(value) -> Optional[float]:
    """
    Convert Danish number format to float.
    Danish uses periods as thousand separators and commas as decimal separators.
    Examples: "72.000" → 72000, "1.234,56" → 1234.56
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if not isinstance(value, str):
        value = str(value)

    value = value.strip()
    if not value:
        return None

    # Remove currency symbols and whitespace
    value = re.sub(r'[kr\s€$]', '', value, flags=re.IGNORECASE)

    # Check if it looks like a number
    if not re.search(r'\d', value):
        return None

    # Handle Danish format: periods as thousands, comma as decimal
    # If there's a comma, it's likely the decimal separator
    if ',' in value:
        # Remove thousand separators (periods)
        value = value.replace('.', '')
        # Convert decimal separator
        value = value.replace(',', '.')
    else:
        # No comma - periods might be thousand separators
        # Count periods to determine format
        period_count = value.count('.')
        if period_count == 1:
            # Could be decimal or thousand separator
            # If 3 digits after period, likely thousand separator
            parts = value.split('.')
            if len(parts[1]) == 3 and len(parts[0]) <= 3:
                value = value.replace('.', '')
        elif period_count > 1:
            # Multiple periods = thousand separators
            value = value.replace('.', '')

    try:
        return float(value)
    except ValueError:
        return None


def find_header_row(rows: list, max_rows: int = 15) -> tuple[int, list]:
    """
    Find the header row by scoring rows based on Danish keyword matches.
    Returns (row_index, header_cells).
    """
    best_score = 0
    best_row_idx = -1
    best_row = None

    for idx, row in enumerate(rows[:max_rows]):
        if not row:
            continue

        # Convert all cells to lowercase strings for matching
        cells = [str(cell).lower().strip() if cell is not None else "" for cell in row]

        # Score based on keyword matches
        score = 0
        for cell in cells:
            for keyword in HEADER_KEYWORDS:
                if keyword in cell:
                    score += 1
                    break  # Only count each cell once

        # Bonus for having multiple non-empty cells
        non_empty = sum(1 for c in cells if c)
        if non_empty >= 3:
            score += 1

        if score > best_score:
            best_score = score
            best_row_idx = idx
            best_row = row

    if best_score < 2:
        raise ParseError("no_header_found")

    return best_row_idx, best_row


def map_columns(headers: list) -> dict:
    """Map Danish headers to standard field names."""
    mapping = {}
    for header in headers:
        if header is None:
            continue
        header_str = str(header).strip()
        header_lower = header_str.lower()

        # Try exact match first
        if header_lower in COLUMN_MAPPING:
            mapping[header_str] = COLUMN_MAPPING[header_lower]
        else:
            # Try partial match
            for danish, standard in COLUMN_MAPPING.items():
                if danish in header_lower or header_lower in danish:
                    mapping[header_str] = standard
                    break

    return mapping


def is_end_row(row: list) -> bool:
    """Check if row indicates end of data (total/sum row or empty)."""
    if not row:
        return True

    # Check if all cells are empty/None
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if not non_empty:
        return True

    # Check for end keywords in first few cells
    for cell in row[:3]:
        if cell is not None:
            cell_lower = str(cell).lower().strip()
            for keyword in END_ROW_KEYWORDS:
                if keyword in cell_lower:
                    return True

    return False


def get_rent_type_from_sheet(sheet_name: str) -> Optional[str]:
    """Determine rent type from sheet name."""
    name_lower = sheet_name.lower()
    if "parkering" in name_lower or "p-plads" in name_lower:
        return "parkering"
    elif "erhverv" in name_lower or "commercial" in name_lower:
        return "erhverv"
    return None


def calculate_summary_stats(rows: list, columns: list, column_mapping: dict) -> tuple[dict, dict]:
    """
    Calculate summary statistics from parsed rows.
    Returns (summary, data_quality) tuple.
    """
    # Reverse mapping: standard name -> column index
    standard_to_idx = {}
    for col_idx, col_name in enumerate(columns):
        if col_name in column_mapping:
            standard_name = column_mapping[col_name]
            standard_to_idx[standard_name] = col_idx

    sqm_idx = standard_to_idx.get("sqm")
    rent_idx = standard_to_idx.get("annual_rent")
    unit_type_idx = standard_to_idx.get("unit_type")
    unit_status_idx = standard_to_idx.get("unit_status")

    # Initialize counters
    total_sqm = 0.0
    total_rent = 0.0
    total_vacant = 0
    units_with_sqm = 0
    units_with_rent = 0
    rent_per_sqm_sum = 0.0
    rent_per_sqm_count = 0

    # Unit type breakdown with count, sqm, rent, and vacant per category
    unit_type_breakdown = {
        "bolig": {"count": 0, "sqm": 0.0, "rent": 0.0, "vacant": 0},
        "erhverv": {"count": 0, "sqm": 0.0, "rent": 0.0, "vacant": 0},
        "parkering": {"count": 0, "sqm": 0.0, "rent": 0.0, "vacant": 0},
        "andet": {"count": 0, "sqm": 0.0, "rent": 0.0, "vacant": 0},
    }

    rows_missing_sqm = []
    rows_missing_rent = []
    rows_suspicious = []

    for row in rows:
        raw = row["raw"]
        row_num = row["row_num"]

        # Extract sqm value
        sqm_value = None
        if sqm_idx is not None and sqm_idx < len(raw):
            sqm_value = convert_danish_number(raw[sqm_idx])
            if sqm_value is not None and sqm_value > 0:
                total_sqm += sqm_value
                units_with_sqm += 1
            else:
                rows_missing_sqm.append(row_num)
        else:
            rows_missing_sqm.append(row_num)

        # Extract rent value
        rent_value = None
        if rent_idx is not None and rent_idx < len(raw):
            rent_value = convert_danish_number(raw[rent_idx])
            if rent_value is not None and rent_value > 0:
                total_rent += rent_value
                units_with_rent += 1
            elif rent_value == 0:
                rows_missing_rent.append(row_num)
                # Check for zero rent with valid sqm
                if sqm_value is not None and sqm_value > 0:
                    rows_suspicious.append({
                        "row_num": row_num,
                        "issue": "Zero rent",
                        "value": 0,
                        "unit": "kr"
                    })
            else:
                rows_missing_rent.append(row_num)
        else:
            rows_missing_rent.append(row_num)

        # Categorize unit type and accumulate sqm/rent per category
        if unit_type_idx is not None and unit_type_idx < len(raw):
            unit_type_value = raw[unit_type_idx]
            category = categorize_unit_type(str(unit_type_value) if unit_type_value else "")
        else:
            category = "andet"

        unit_type_breakdown[category]["count"] += 1
        if sqm_value is not None and sqm_value > 0:
            unit_type_breakdown[category]["sqm"] += sqm_value
        if rent_value is not None and rent_value > 0:
            unit_type_breakdown[category]["rent"] += rent_value

        # Check vacancy status
        if unit_status_idx is not None and unit_status_idx < len(raw):
            status_value = raw[unit_status_idx]
            if is_unit_vacant(str(status_value) if status_value else ""):
                unit_type_breakdown[category]["vacant"] += 1
                total_vacant += 1

        # Calculate rent per sqm for this row
        if sqm_value is not None and sqm_value > 0 and rent_value is not None and rent_value > 0:
            row_rent_per_sqm = rent_value / sqm_value
            rent_per_sqm_sum += row_rent_per_sqm
            rent_per_sqm_count += 1

    # Calculate average rent per sqm
    avg_rent_per_sqm = 0.0
    if rent_per_sqm_count > 0:
        avg_rent_per_sqm = rent_per_sqm_sum / rent_per_sqm_count

    # Find unmapped columns
    unmapped_columns = []
    for col in columns:
        if col and col.strip() and col not in column_mapping:
            unmapped_columns.append(col)

    # Round sqm and rent values in breakdown
    for category in unit_type_breakdown:
        unit_type_breakdown[category]["sqm"] = round(unit_type_breakdown[category]["sqm"], 2)
        unit_type_breakdown[category]["rent"] = round(unit_type_breakdown[category]["rent"], 2)

    summary = {
        "total_units": len(rows),
        "total_sqm": round(total_sqm, 2),
        "total_annual_rent": round(total_rent, 2),
        "avg_rent_per_sqm": round(avg_rent_per_sqm, 2),
        "units_with_rent": units_with_rent,
        "units_with_sqm": units_with_sqm,
        "total_vacant": total_vacant,
        "unit_type_breakdown": unit_type_breakdown,
    }

    data_quality = {
        "rows_missing_sqm": rows_missing_sqm,
        "rows_missing_rent": rows_missing_rent,
        "rows_suspicious": rows_suspicious,
        "unmapped_columns": unmapped_columns,
    }

    return summary, data_quality


def parse_excel(file_path: str) -> dict:
    """Parse Excel file (.xlsx or .xls)."""
    ext = Path(file_path).suffix.lower()

    result = {
        "filename": Path(file_path).name,
        "file_type": "excel",
        "source_info": {
            "sheets_found": [],
            "sheets_used": [],
        },
        "header_row": None,
        "columns": [],
        "column_mapping": {},
        "rows": [],
        "total_rows": 0,
        "parse_warnings": [],
        "confidence": "high",
    }

    try:
        if ext == ".xlsx":
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
        else:  # .xls
            workbook = xlrd.open_workbook(file_path)
            sheet_names = workbook.sheet_names()
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypted" in error_msg:
            raise ParseError("password_protected")
        raise ParseError("corrupted_file", f"Could not open file: {str(e)}")

    result["source_info"]["sheets_found"] = sheet_names

    all_rows = []
    header_found = False

    for sheet_name in sheet_names:
        try:
            if ext == ".xlsx":
                sheet = workbook[sheet_name]
                rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
            else:
                sheet = workbook.sheet_by_name(sheet_name)
                rows = [sheet.row_values(i) for i in range(sheet.nrows)]

            if len(rows) < 3:
                result["parse_warnings"].append(f"Sheet '{sheet_name}' skipped - insufficient rows")
                continue

            # Find header row if not found yet
            if not header_found:
                try:
                    header_idx, header_row = find_header_row(rows)
                    result["header_row"] = header_idx + 1  # 1-indexed
                    result["columns"] = [str(c) if c is not None else "" for c in header_row]
                    result["column_mapping"] = map_columns(header_row)
                    header_found = True
                except ParseError:
                    result["parse_warnings"].append(f"Sheet '{sheet_name}' skipped - no header found")
                    continue
            else:
                # For subsequent sheets, try to find matching header
                try:
                    header_idx, _ = find_header_row(rows)
                except ParseError:
                    result["parse_warnings"].append(f"Sheet '{sheet_name}' skipped - no header found")
                    continue

            # Extract data rows
            rent_type = get_rent_type_from_sheet(sheet_name)
            data_rows = rows[header_idx + 1:]
            sheet_row_count = 0

            for row_offset, row in enumerate(data_rows):
                if is_end_row(row):
                    break

                # Clean row data
                raw_data = [str(c) if c is not None else "" for c in row]

                # Skip rows that are mostly empty
                non_empty = sum(1 for c in raw_data if c.strip())
                if non_empty < 2:
                    continue

                all_rows.append({
                    "raw": raw_data,
                    "source": sheet_name,
                    "row_num": header_idx + row_offset + 2,  # 1-indexed, after header
                    "rent_type": rent_type,
                })
                sheet_row_count += 1

            if sheet_row_count > 0:
                result["source_info"]["sheets_used"].append(sheet_name)
            else:
                result["parse_warnings"].append(f"Sheet '{sheet_name}' skipped - no data rows")

        except Exception as e:
            result["parse_warnings"].append(f"Sheet '{sheet_name}' error: {str(e)}")
            result["confidence"] = "medium"

    if ext == ".xlsx":
        workbook.close()

    if not header_found:
        raise ParseError("no_header_found")

    if not all_rows:
        raise ParseError("no_data_found")

    result["rows"] = all_rows
    result["total_rows"] = len(all_rows)

    # Calculate summary statistics
    summary, data_quality = calculate_summary_stats(
        all_rows, result["columns"], result["column_mapping"]
    )
    result["summary"] = summary
    result["data_quality"] = data_quality

    return result


def parse_pdf(file_path: str) -> dict:
    """Parse PDF file with table extraction and OCR fallback."""
    result = {
        "filename": Path(file_path).name,
        "file_type": "pdf",
        "source_info": {
            "pages": 0,
            "tables_found": 0,
            "ocr_used": False,
        },
        "header_row": None,
        "columns": [],
        "column_mapping": {},
        "rows": [],
        "total_rows": 0,
        "parse_warnings": [],
        "confidence": "high",
    }

    all_tables = []

    try:
        with pdfplumber.open(file_path) as pdf:
            result["source_info"]["pages"] = len(pdf.pages)

            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()

                for table in tables:
                    if table and len(table) > 1:
                        all_tables.append({
                            "page": page_num,
                            "rows": table,
                        })

            result["source_info"]["tables_found"] = len(all_tables)
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypted" in error_msg:
            raise ParseError("password_protected")
        raise ParseError("corrupted_file", f"Could not open PDF: {str(e)}")

    # If no tables found, try OCR
    if not all_tables:
        result["parse_warnings"].append("No tables found with pdfplumber, attempting OCR")
        all_tables = extract_tables_with_ocr(file_path, result)
        result["source_info"]["ocr_used"] = True
        result["confidence"] = "low"

    if not all_tables:
        raise ParseError("no_data_found", "No tables found in PDF")

    # Merge all tables and find header
    all_rows = []
    header_found = False
    seen_headers = set()

    for table_info in all_tables:
        rows = table_info["rows"]
        page_num = table_info["page"]

        if not header_found:
            try:
                header_idx, header_row = find_header_row(rows)
                result["header_row"] = header_idx + 1
                result["columns"] = [str(c) if c is not None else "" for c in header_row]
                result["column_mapping"] = map_columns(header_row)
                header_found = True

                # Create header signature for deduplication
                header_sig = tuple(str(c).lower().strip() for c in header_row if c)
                seen_headers.add(header_sig)

                # Process rows after header
                data_rows = rows[header_idx + 1:]
            except ParseError:
                continue
        else:
            # Check if this table starts with a repeated header
            if rows:
                first_row = rows[0]
                first_row_sig = tuple(str(c).lower().strip() for c in first_row if c)

                if first_row_sig in seen_headers:
                    # Skip repeated header
                    data_rows = rows[1:]
                else:
                    data_rows = rows

        # Extract data rows
        for row_offset, row in enumerate(data_rows):
            if is_end_row(row):
                continue

            raw_data = [str(c) if c is not None else "" for c in row]

            # Skip mostly empty rows
            non_empty = sum(1 for c in raw_data if c.strip())
            if non_empty < 2:
                continue

            # Skip if this looks like a repeated header
            row_sig = tuple(c.lower().strip() for c in raw_data if c.strip())
            if row_sig in seen_headers:
                continue

            all_rows.append({
                "raw": raw_data,
                "source": f"page_{page_num}",
                "row_num": row_offset + 1,
                "rent_type": None,
            })

    if not header_found:
        raise ParseError("no_header_found")

    if not all_rows:
        raise ParseError("no_data_found")

    result["rows"] = all_rows
    result["total_rows"] = len(all_rows)

    # Calculate summary statistics
    summary, data_quality = calculate_summary_stats(
        all_rows, result["columns"], result["column_mapping"]
    )
    result["summary"] = summary
    result["data_quality"] = data_quality

    return result


def extract_tables_with_ocr(file_path: str, result: dict) -> list:
    """Extract tables from scanned PDF using OCR."""
    tables = []

    try:
        images = convert_from_path(file_path, dpi=300)
    except Exception as e:
        result["parse_warnings"].append(f"OCR image conversion failed: {str(e)}")
        raise ParseError("ocr_failed", str(e))

    for page_num, image in enumerate(images, 1):
        try:
            # Use pytesseract to extract text
            text = pytesseract.image_to_string(image, lang='dan+eng')

            # Split into lines and parse as table
            lines = text.strip().split('\n')
            rows = []

            for line in lines:
                if line.strip():
                    # Split by multiple spaces or tabs
                    cells = re.split(r'\s{2,}|\t', line)
                    cells = [c.strip() for c in cells if c.strip()]
                    if cells:
                        rows.append(cells)

            if rows:
                tables.append({
                    "page": page_num,
                    "rows": rows,
                })
        except Exception as e:
            result["parse_warnings"].append(f"OCR failed for page {page_num}: {str(e)}")

    return tables


def parse_rent_roll(file_path: str) -> dict:
    """
    Main entry point for parsing rent roll files.
    Detects file type and routes to appropriate parser.
    """
    if not os.path.exists(file_path):
        raise ParseError("file_not_found", f"File not found: {file_path}")

    file_type = detect_file_type(file_path)

    if file_type == "excel":
        return parse_excel(file_path)
    elif file_type == "pdf":
        return parse_pdf(file_path)
    else:
        raise ParseError("invalid_file_type")
